#!/usr/bin/python3
from openpyxl import load_workbook 

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

total = {}
r = 1
for row in ws:
	sum = 0.0
	if r != 1:
		sum += ws.cell(row = r, column = 3).value * 0.3
		sum += ws.cell(row = r, column = 4).value * 0.35
		sum += ws.cell(row = r, column = 5).value * 0.34
		sum += ws.cell(row = r, column = 6).value
		ws.cell(row = r, column = 7, value = sum)
		total[ws.cell(row = r, column = 7).value] = r
	r += 1

print(total)
total_sort = sorted(total)
total_sort.reverse()
print(total_sort)

aCutIdx = int(len(total_sort) * 0.3) - 1
print(aCutIdx)
#aCutIdx = int(len(total_sort) * 0.3) - 1
aCut = total_sort[aCutIdx]
print("aCut is " , aCut, ", ", aCutIdx)
bCutIdx = int(len(total_sort) * 0.7) - 1
bCut = total_sort[bCutIdx]
print("bCut is ", bCut, ", ", bCutIdx)

aPCutIdx = int(aCutIdx * 0.5)
aPlusCut = total_sort[aPCutIdx]
bPCutIdx = int((bCutIdx - aCutIdx) * 0.5) + aCutIdx
bPlusCut = total_sort[bPCutIdx]
cPCutIdx = int((len(total_sort) - bCutIdx) * 0.5) + bCutIdx
cPlusCut = total_sort[cPCutIdx]
print(aPlusCut, ", ", bPlusCut, ", ", cPlusCut)

for i in range(0, len(total_sort)):
	if i <= aPCutIdx:
		r = total[total_sort[i]]
		print(r)
		ws.cell(row = r, column = 8, value = "A+")
	elif i <= aCutIdx:
		r = total[total_sort[i]]
		print(r)
		ws.cell(row = r, column = 8, value = "A0")
	elif i <= bPCutIdx:
		r = total[total_sort[i]]
		ws.cell(row = r, column = 8, value = "B+")
	elif i <= bCutIdx:
		r = total[total_sort[i]]
		ws.cell(row = r, column = 8, value = "B0")
	elif i <= cPCutIdx:
		r = total[total_sort[i]]
		ws.cell(row = r, column = 8, value = "C+")
	else:
		r = total[total_sort[i]]
		ws.cell(row = r, column = 8, value = "C0")
		
wb.save("student.xlsx")
		

