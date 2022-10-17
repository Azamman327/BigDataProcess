#!/usr/bin/python3
from openpyxl import load_workbook 

wb = load_workbook(filename = 'student.xlsx')
ws = wb['Sheet1']

r = 1
for row in ws:
	sum = 0.0
	if r != 1:
		sum += ws.cell(row = r, column = 3).value * 0.3
		sum += ws.cell(row = r, column = 4).value * 0.35
		sum += ws.cell(row = r, column = 5).value * 0.34
		sum += ws.cell(row = r, column = 6).value
		ws.cell(row = r, column = 7, value = sum)
	r += 1

wb.save("student.xlsx")
		

